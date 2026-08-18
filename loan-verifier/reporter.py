"""
reporter.py — Generates the Excel verification reports.

build_report()          — interest verifier (verify_loans.py): Summary + Details
                           sheets for nominal/effective interest only.
build_payments_report() — payments verifier (verify_payments.py): Payments
                           Summary/Details + Deals Reconciliation sheets.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import REPORT_FOLDER, REPORT_FILENAME_PREFIX, PAYMENTS_REPORT_FILENAME_PREFIX
from models import VerificationRow, PaymentCheckRow, DealReconciliationRow

logger = logging.getLogger(__name__)

# Colours
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_report(rows: List[VerificationRow]) -> str:
    """
    Write the interest-verification results (nominal + effective) to a
    timestamped Excel file. Returns the full path of the created file.
    """
    if not rows:
        logger.warning("No verification rows to report.")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"{REPORT_FILENAME_PREFIX}_{timestamp}.xlsx"
    filepath = str(Path(REPORT_FOLDER) / filename)

    wb = Workbook()
    _write_summary_sheet(wb, rows)
    _write_details_sheet(wb, rows)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    logger.info(f"Report saved: {filepath}")
    return filepath


def build_payments_report(
    payment_rows: List[PaymentCheckRow],
    deal_rows: Optional[List[DealReconciliationRow]] = None,
) -> str:
    """
    Write the payments-table verification results (Phase 1 self-consistency
    checks + Phase 2 deals reconciliation) to a separate timestamped Excel
    file. Returns the full path of the created file.
    """
    if not payment_rows:
        logger.warning("No payment rows to report.")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"{PAYMENTS_REPORT_FILENAME_PREFIX}_{timestamp}.xlsx"
    filepath = str(Path(REPORT_FOLDER) / filename)

    wb = Workbook()
    _write_payments_summary_sheet(wb, payment_rows)
    _write_payments_details_sheet(wb, payment_rows)
    if deal_rows is not None:
        _write_deals_reconciliation_sheet(wb, deal_rows)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    logger.info(f"Payments report saved: {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(wb: Workbook, rows: List[VerificationRow]) -> None:
    ws = wb.create_sheet("Summary")

    # Aggregate by contract
    contracts: dict = {}
    for r in rows:
        key = (r.contract_id, r.contract_num, r.client_id, r.client_name)
        if key not in contracts:
            contracts[key] = {
                "total": 0, "nom_mismatch": 0, "eff_mismatch": 0,
                "nom_missing": 0, "eff_missing": 0,
            }
        contracts[key]["total"] += 1
        if not r.has_stored_record:
            if r.interest_type == "Nominal":
                contracts[key]["nom_missing"] += 1
            else:
                contracts[key]["eff_missing"] += 1
        elif not r.is_match:
            if r.interest_type == "Nominal":
                contracts[key]["nom_mismatch"] += 1
            else:
                contracts[key]["eff_mismatch"] += 1

    headers = [
        "Contract ID", "Contract #", "Client ID", "Client",
        "Total Days", "Nominal Mismatches", "Effective Mismatches",
        "Nominal Missing", "Effective Missing", "Status",
    ]
    _write_header_row(ws, headers)

    for row_idx, ((cid, cnum, clid, cname), stats) in enumerate(
        sorted(contracts.items()), start=2
    ):
        total_mm      = stats["nom_mismatch"] + stats["eff_mismatch"]
        total_missing = stats["nom_missing"]  + stats["eff_missing"]
        if total_mm == 0 and total_missing == 0:
            status = "OK"
            fill = GREEN_FILL
        elif total_mm > 0:
            status = f"{total_mm} MISMATCH(ES)"
            fill = RED_FILL
        else:
            status = f"{total_missing} MISSING"
            fill = ORANGE_FILL
        data = [
            cid, cnum, clid, cname,
            stats["total"] // 2,  # days (2 rows per day: nominal + effective)
            stats["nom_mismatch"], stats["eff_mismatch"],
            stats["nom_missing"], stats["eff_missing"],
            status,
        ]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Details sheet
# ---------------------------------------------------------------------------

def _write_details_sheet(wb: Workbook, rows: List[VerificationRow]) -> None:
    ws = wb.create_sheet("Details")

    headers = [
        "Contract ID", "Contract #", "Client ID", "Client", "Date", "Interest Type",
        "Opening Balance (AMD)", "Daily Rate (%)",
        "Calculated Interest", "Stored Interest", "Difference", "Stored?", "Match?",
    ]
    _write_header_row(ws, headers)

    for row_idx, r in enumerate(
        sorted(rows, key=lambda x: (x.contract_id, x.record_date, x.interest_type)),
        start=2,
    ):
        if not r.has_stored_record:
            fill = ORANGE_FILL
            match_label = "MISSING"
        elif r.is_match:
            fill = GREEN_FILL
            match_label = "YES"
        else:
            fill = RED_FILL
            match_label = "NO"

        data = [
            r.contract_id,
            r.contract_num,
            r.client_id,
            r.client_name,
            r.record_date.isoformat() if r.record_date else "",
            r.interest_type,
            r.opening_balance,
            r.daily_rate_pct,
            r.calculated_interest,
            r.stored_interest if r.has_stored_record else "—",
            r.difference if r.has_stored_record else "—",
            "YES" if r.has_stored_record else "NO",
            match_label,
        ]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Format numeric columns with 2 decimal places
            if col_idx in (7, 9, 10, 11) and isinstance(value, float):
                cell.number_format = "#,##0.00"
            if col_idx == 8:
                cell.number_format = "0.000000"

    # Freeze header row
    ws.freeze_panes = "A2"
    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Payments Summary sheet
# ---------------------------------------------------------------------------

def _write_payments_summary_sheet(wb: Workbook, rows: List[PaymentCheckRow]) -> None:
    ws = wb.create_sheet("Payments Summary")

    contracts: dict = {}
    for r in rows:
        key = (r.contract_id, r.contract_num, r.client_id, r.client_name)
        stats = contracts.setdefault(key, {
            "total": 0, "amount_mismatch": 0, "status_issue": 0, "remaining_increase": 0,
        })
        stats["total"] += 1
        if not r.amount_match:
            stats["amount_mismatch"] += 1
        if not r.status_ok:
            stats["status_issue"] += 1
        if r.remaining_increased:
            stats["remaining_increase"] += 1

    headers = [
        "Contract ID", "Contract #", "Client ID", "Client",
        "Total Payments", "Amount Mismatches", "Status Issues",
        "Remaining Increases", "Status",
    ]
    _write_header_row(ws, headers)

    for row_idx, ((cid, cnum, clid, cname), stats) in enumerate(
        sorted(contracts.items()), start=2
    ):
        total_issues = stats["amount_mismatch"] + stats["status_issue"] + stats["remaining_increase"]
        if total_issues == 0:
            status = "OK"
            fill = GREEN_FILL
        else:
            status = f"{total_issues} ISSUE(S)"
            fill = RED_FILL
        data = [
            cid, cnum, clid, cname,
            stats["total"], stats["amount_mismatch"], stats["status_issue"],
            stats["remaining_increase"], status,
        ]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Payments Details sheet
# ---------------------------------------------------------------------------

def _write_payments_details_sheet(wb: Workbook, rows: List[PaymentCheckRow]) -> None:
    ws = wb.create_sheet("Payments Details")

    headers = [
        "Contract ID", "Contract #", "Client ID", "Client", "Payment ID", "Date",
        "Type", "Status", "Amount", "Principal Payment", "Interest Payment",
        "Calc Amount", "Amount Diff", "Amount OK?", "Paid", "Status OK?",
        "Remaining", "Prev Remaining", "Remaining Increased?", "Match?",
    ]
    _write_header_row(ws, headers)

    for row_idx, r in enumerate(
        sorted(rows, key=lambda x: (x.contract_id, x.payment_date, x.payment_id)),
        start=2,
    ):
        if r.is_match:
            fill = GREEN_FILL
            match_label = "YES"
        elif r.remaining_increased and r.amount_match and r.status_ok:
            fill = ORANGE_FILL
            match_label = "REVIEW"
        else:
            fill = RED_FILL
            match_label = "NO"

        data = [
            r.contract_id,
            r.contract_num,
            r.client_id,
            r.client_name,
            r.payment_id,
            r.payment_date.isoformat() if r.payment_date else "",
            r.payment_type,
            r.status,
            r.amount,
            r.principal_payment,
            r.interest_payment,
            r.calc_amount,
            r.amount_diff,
            "YES" if r.amount_match else "NO",
            r.paid if r.paid is not None else "—",
            "YES" if r.status_ok else "NO",
            r.remaining,
            r.prev_remaining if r.prev_remaining is not None else "—",
            "YES" if r.remaining_increased else "NO",
            match_label,
        ]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Format numeric columns with 2 decimal places
            if col_idx in (9, 10, 11, 12, 13, 15, 17, 18) and isinstance(value, float):
                cell.number_format = "#,##0.00"

    # Freeze header row
    ws.freeze_panes = "A2"
    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Deals Reconciliation sheet
# ---------------------------------------------------------------------------

def _write_deals_reconciliation_sheet(wb: Workbook, rows: List[DealReconciliationRow]) -> None:
    ws = wb.create_sheet("Deals Reconciliation")

    headers = [
        "Contract ID", "Contract #", "Client ID", "Client", "Seq #",
        "Payment ID", "Payment Date (scheduled)", "Deal ID", "Deal Date (actual)",
        "Original Interest", "Deal Interest", "Interest Diff", "Interest Issue?",
        "Original Principal", "Deal Amount", "Deal Penalty", "Principal Diff (review only)",
        "Penalty Payment ID", "Penalty Match?", "Unpaired?",
    ]
    _write_header_row(ws, headers)

    for row_idx, r in enumerate(
        sorted(rows, key=lambda x: (x.contract_id, x.sequence_index)),
        start=2,
    ):
        if r.unpaired:
            fill = ORANGE_FILL
        elif r.interest_flag or not r.penalty_match:
            fill = RED_FILL
        else:
            fill = GREEN_FILL

        data = [
            r.contract_id,
            r.contract_num,
            r.client_id,
            r.client_name,
            r.sequence_index,
            r.payment_id if r.payment_id is not None else "—",
            r.payment_date.isoformat() if r.payment_date else "—",
            r.deal_id if r.deal_id is not None else "—",
            r.deal_date.isoformat() if r.deal_date else "—",
            r.original_interest_payment if r.original_interest_payment is not None else "—",
            r.deal_interest_amount if r.deal_interest_amount is not None else "—",
            r.interest_diff if r.interest_diff is not None else "—",
            "YES" if r.interest_flag else "NO",
            r.original_principal_payment if r.original_principal_payment is not None else "—",
            r.deal_amount if r.deal_amount is not None else "—",
            r.deal_penalty if r.deal_penalty is not None else "—",
            r.principal_diff if r.principal_diff is not None else "—",
            r.penalty_payment_id if r.penalty_payment_id is not None else "—",
            "YES" if r.penalty_match else "NO",
            "YES" if r.unpaired else "NO",
        ]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_idx in (10, 11, 12, 14, 15, 16, 17) and isinstance(value, float):
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_header_row(ws, headers: List[str]) -> None:
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30


def _auto_column_widths(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

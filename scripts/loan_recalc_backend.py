#!/usr/bin/env python3
from __future__ import annotations

import argparse
import ast
import calendar
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class ScheduleRow:
    idx: int
    from_date: date
    due_date: date
    days: int
    opening: float
    interest_due: float
    principal_due: float
    payment_due: float
    penalty_due: float = 0.0
    paid_interest: float = 0.0
    paid_principal: float = 0.0
    paid_penalty: float = 0.0
    paid_total: float = 0.0
    payment_date: date | None = None
    status: str = "Pending"

    @property
    def remaining_interest(self) -> float:
        return max(0.0, self.interest_due - self.paid_interest)

    @property
    def remaining_principal(self) -> float:
        return max(0.0, self.principal_due - self.paid_principal)

    @property
    def closing_balance(self) -> float:
        return max(0.0, self.opening - self.paid_principal)


@dataclass
class PaymentEntry:
    pay_date: date
    installment: Optional[int]
    amount: float


def to_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    raise ValueError(f"Unsupported date value: {value!r}")


def add_months(src: date, months: int) -> date:
    month = src.month - 1 + months
    year = src.year + month // 12
    month = month % 12 + 1
    day = min(src.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def next_working_day(raw: date, holidays: set[date]) -> date:
    d = raw
    while d.weekday() >= 5 or d in holidays:
        d += timedelta(days=1)
    return d


def pmt(rate: float, nper: int, pv: float) -> float:
    if abs(rate) < 1e-12:
        return pv / nper
    return (rate * pv) / (1 - (1 + rate) ** (-nper))


def recalc_future_interest(rows: List[ScheduleRow], from_index: int, outstanding_principal: float, daily_rate_percent: float) -> None:
    balance = max(0.0, outstanding_principal)
    for i in range(from_index, len(rows)):
        row = rows[i]
        row.opening = balance
        row.interest_due = balance * row.days * daily_rate_percent / 100.0
        row.payment_due = row.principal_due + row.interest_due
        if row.paid_principal > row.principal_due:
            row.paid_principal = row.principal_due
        if row.paid_interest > row.interest_due:
            row.paid_interest = row.interest_due
        balance = max(0.0, balance - row.principal_due)


def enforce_annuity_opening_chain(rows: List[ScheduleRow], daily_rate_percent: float) -> None:
    """
    Keep schedule in annuity-style chain:
    opening[i] = opening[i-1] - principal_due[i-1]
    interest_due[i] = opening[i] * days[i] * daily_rate
    payment_due[i] = principal_due[i] + interest_due[i]
    """
    if not rows:
        return
    for i in range(1, len(rows)):
        prev = rows[i - 1]
        cur = rows[i]
        cur.opening = max(0.0, prev.opening - prev.principal_due)
        cur.interest_due = cur.opening * cur.days * daily_rate_percent / 100.0
        cur.payment_due = cur.principal_due + cur.interest_due


def apply_extra_to_future_principal(rows: List[ScheduleRow], start_index: int, extra: float, daily_rate_percent: float) -> float:
    remain = max(0.0, extra)
    if remain <= 0:
        return 0.0
    for i in range(start_index, len(rows)):
        if remain <= 0:
            break
        row = rows[i]
        reducible = max(0.0, row.principal_due - row.paid_principal)
        reduction = min(remain, reducible)
        if reduction > 0:
            row.principal_due -= reduction
            row.payment_due = row.principal_due + row.interest_due
            remain -= reduction
    if start_index < len(rows):
        prior_balance = rows[start_index - 1].closing_balance if start_index > 0 else rows[0].opening
        recalc_future_interest(rows, start_index, prior_balance, daily_rate_percent)
    return remain


def build_base_schedule(
    loan_amount: float,
    daily_rate_percent: float,
    fee_annual_percent: float,
    start_date: date,
    first_payment_date: date,
    months: int,
    holidays: set[date],
) -> List[ScheduleRow]:
    monthly_rate = daily_rate_percent * 365 / 100 / 12 + fee_annual_percent / 100 / 12
    emi = pmt(monthly_rate, months, loan_amount)

    rows: List[ScheduleRow] = []
    balance = loan_amount
    base_date = start_date

    for i in range(1, months + 1):
        prev_raw = add_months(base_date, i - 1)
        due_raw = add_months(first_payment_date, i - 1)
        from_date = next_working_day(prev_raw, holidays)
        due_date = next_working_day(due_raw, holidays)
        days = max(1, (due_date - from_date).days)
        interest_due = balance * days * daily_rate_percent / 100.0
        principal_due = emi - interest_due
        if i == months:
            principal_due = balance
        principal_due = max(0.0, principal_due)
        payment_due = principal_due + interest_due
        rows.append(
            ScheduleRow(
                idx=i,
                from_date=from_date,
                due_date=due_date,
                days=days,
                opening=balance,
                interest_due=interest_due,
                principal_due=principal_due,
                payment_due=payment_due,
            )
        )
        balance = max(0.0, balance - principal_due)
    return rows


def try_early_split(row: ScheduleRow, contract_outstanding: float, daily_rate_percent: float, cash: float, pay_date: date):
    if pay_date >= row.due_date:
        return None
    elapsed = max(1, (pay_date - row.from_date).days)
    future_days = (row.due_date - pay_date).days
    if future_days < 1:
        return None
    p = max(0.0, contract_outstanding)
    if p <= 0 or cash <= 0:
        return None
    rate = daily_rate_percent / 100.0
    past_interest = p * elapsed * rate
    k_future = future_days * rate
    denom = 1 - k_future
    if abs(denom) < 1e-12:
        return None
    x = (cash - past_interest - p * k_future) / denom
    if x < 0:
        return None
    x = min(x, p, cash)
    future_interest = max(0.0, (max(0.0, p - x) * future_days * rate))
    paid_interest = past_interest + future_interest
    principal_for_line = min(x, row.remaining_principal)
    remain_cash = max(0.0, cash - paid_interest - principal_for_line)
    return paid_interest, principal_for_line, remain_cash


def _apply_cash_to_row(rows: List[ScheduleRow], row_idx_1based: int, pay_date: date, cash: float, daily_rate_percent: float) -> float:
    row = rows[row_idx_1based - 1]
    delay_days = max(0, (pay_date - row.due_date).days)
    outstanding_due = row.remaining_interest + row.remaining_principal
    if delay_days > 0:
        row.penalty_due += outstanding_due * delay_days * daily_rate_percent / 100.0

    penalty_pay = min(cash, max(0.0, row.penalty_due - row.paid_penalty))
    row.paid_penalty += penalty_pay
    cash -= penalty_pay

    if cash + 10 >= outstanding_due:
        contract_outstanding = row.opening
        split = try_early_split(row, contract_outstanding, daily_rate_percent, cash, pay_date)
        if split is not None:
            paid_interest, paid_principal, cash = split
            row.paid_interest += min(row.remaining_interest, paid_interest)
            row.paid_principal += min(row.remaining_principal, paid_principal)
        else:
            int_pay = min(cash, row.remaining_interest)
            row.paid_interest += int_pay
            cash -= int_pay
            princ_pay = min(cash, row.remaining_principal)
            row.paid_principal += princ_pay
            cash -= princ_pay
    else:
        int_pay = min(cash, row.remaining_interest)
        row.paid_interest += int_pay
        cash -= int_pay
        princ_pay = min(cash, row.remaining_principal)
        row.paid_principal += princ_pay
        cash -= princ_pay
    return cash


def apply_payments(rows: List[ScheduleRow], payments: List[PaymentEntry], daily_rate_percent: float) -> None:
    payments_sorted = sorted(payments, key=lambda p: (p.pay_date, p.installment or 0))
    for entry in payments_sorted:
        cash = max(0.0, entry.amount)
        if cash <= 0:
            continue

        # If installment is given, use it. Otherwise allocate backend-style:
        # oldest eligible unpaid schedule rows first.
        if entry.installment is not None:
            target_indices = [entry.installment]
        else:
            target_indices = [
                r.idx
                for r in rows
                if r.from_date < entry.pay_date and (r.remaining_interest + r.remaining_principal + max(0.0, r.penalty_due - r.paid_penalty)) > 1e-8
            ]
            if not target_indices:
                target_indices = [r.idx for r in rows if (r.remaining_interest + r.remaining_principal) > 1e-8]

        total_used = 0.0
        for idx in target_indices:
            if cash <= 1e-8:
                break
            before = cash
            cash = _apply_cash_to_row(rows, idx, entry.pay_date, cash, daily_rate_percent)
            used = max(0.0, before - cash)
            if used > 0:
                row = rows[idx - 1]
                row.paid_total += used
                row.payment_date = entry.pay_date
                total_used += used

        if cash > 0:
            next_unpaid_idx = next((r.idx for r in rows if (r.remaining_principal + r.remaining_interest) > 1e-8), len(rows))
            _ = apply_extra_to_future_principal(rows, next_unpaid_idx, cash, daily_rate_percent)
            if rows:
                rows[min(next_unpaid_idx, len(rows)) - 1].paid_total += cash

    for row in rows:
        total_due = row.interest_due + row.principal_due + row.penalty_due
        total_paid = row.paid_interest + row.paid_principal + row.paid_penalty
        if total_paid + 1e-8 >= total_due and total_due > 0:
            row.status = "Paid"
        elif total_paid > 0:
            row.status = "Partial"
        elif date.today() > row.due_date:
            row.status = "Overdue"
        else:
            row.status = "Pending"


def read_holidays(ws) -> set[date]:
    holidays: set[date] = set()
    for r in range(2, ws.max_row + 1):
        value = ws.cell(r, 1).value
        if value is None:
            continue
        holidays.add(to_date(value))
    return holidays


def parse_amount_cell(value, emi_value: float) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        expr = value.strip()
        if not expr:
            return None
        if expr.startswith("="):
            expr = expr[1:]
            expr = expr.replace("EMI", str(emi_value))
            allowed_nodes = (
                ast.Expression,
                ast.BinOp,
                ast.UnaryOp,
                ast.Num,
                ast.Constant,
                ast.Add,
                ast.Sub,
                ast.Mult,
                ast.Div,
                ast.Pow,
                ast.USub,
                ast.UAdd,
                ast.Load,
            )
            node = ast.parse(expr, mode="eval")
            if not all(isinstance(n, allowed_nodes) for n in ast.walk(node)):
                return None
            return float(eval(compile(node, "<amount_formula>", "eval"), {"__builtins__": {}}, {}))
        try:
            return float(expr)
        except ValueError:
            return None
    return None


def read_payments(ws, emi_value: float) -> List[PaymentEntry]:
    rows: List[PaymentEntry] = []
    for r in range(3, ws.max_row + 1):
        pay_date = ws.cell(r, 1).value
        inst = ws.cell(r, 2).value
        amount_raw = ws.cell(r, 3).value
        amount = parse_amount_cell(amount_raw, emi_value)
        if pay_date is None or amount is None:
            continue
        inst_val: Optional[int] = None
        if inst not in (None, ""):
            inst_val = int(inst)
        rows.append(PaymentEntry(pay_date=to_date(pay_date), installment=inst_val, amount=float(amount)))
    rows.sort(key=lambda x: (x.pay_date, x.installment or 0))
    return rows


def write_output_sheet(wb, rows: List[ScheduleRow]) -> None:
    if "Backend_Recalc" in wb.sheetnames:
        del wb["Backend_Recalc"]
    ws = wb.create_sheet("Backend_Recalc")
    headers = [
        "Installment",
        "From Date",
        "Due Date",
        "Days",
        "Opening",
        "Interest Due",
        "Principal Due",
        "Penalty Due",
        "Paid Interest",
        "Paid Principal",
        "Paid Penalty",
        "Paid Total",
        "Closing",
        "Status",
        "Last Payment Date",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row.idx,
                row.from_date,
                row.due_date,
                row.days,
                round(row.opening, 10),
                round(row.interest_due, 10),
                round(row.principal_due, 10),
                round(row.penalty_due, 10),
                round(row.paid_interest, 10),
                round(row.paid_principal, 10),
                round(row.paid_penalty, 10),
                round(row.paid_total, 10),
                round(row.closing_balance, 10),
                row.status,
                row.payment_date,
            ]
        )


def write_schedule_sheet(wb, rows: List[ScheduleRow]) -> None:
    ws = wb["Schedule"]
    start_row = 3
    for row in rows:
        r = start_row + row.idx - 1
        unpaid_carried = max(
            0.0,
            row.interest_due
            + row.principal_due
            + row.penalty_due
            - row.paid_interest
            - row.paid_principal
            - row.paid_penalty,
        )
        ws.cell(r, 1).value = row.idx
        ws.cell(r, 2).value = row.from_date
        ws.cell(r, 3).value = row.due_date
        ws.cell(r, 4).value = row.days
        ws.cell(r, 5).value = round(row.opening, 10)
        ws.cell(r, 6).value = round(row.interest_due, 10)
        ws.cell(r, 7).value = round(row.principal_due, 10)
        ws.cell(r, 8).value = round(row.paid_total, 10)
        ws.cell(r, 9).value = row.payment_date
        ws.cell(r, 10).value = round(row.paid_interest, 10)
        ws.cell(r, 11).value = round(row.paid_principal, 10)
        ws.cell(r, 12).value = round(unpaid_carried, 10)
        ws.cell(r, 13).value = round(row.closing_balance, 10)
        ws.cell(r, 14).value = row.status

    # Clear trailing old rows for a cleaner schedule view.
    for r in range(start_row + len(rows), ws.max_row + 1):
        for c in range(1, 15):
            ws.cell(r, c).value = None


def main() -> None:
    parser = argparse.ArgumentParser(description="Backend-like amortized recalculation for Loan_Calculator workbook.")
    parser.add_argument("--input", default="Loan_Calculator.xlsx", help="Input workbook path")
    parser.add_argument("--output", default="Loan_Calculator_recalculated.xlsx", help="Output workbook path")
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        raise FileNotFoundError(f"Workbook not found: {in_path}")

    wb = load_workbook(in_path)
    inputs = wb["Inputs"]
    loan_amount = float(inputs["B4"].value)
    daily_rate_percent = float(inputs["B5"].value)
    fee_annual_percent = float(inputs["B6"].value or 0.0)
    start_date = to_date(inputs["B8"].value)
    first_payment_date = to_date(inputs["B9"].value)
    months = int(inputs["B10"].value)
    holidays = read_holidays(wb["Holidays"])

    calc_monthly_rate = daily_rate_percent * 365 / 100 / 12 + fee_annual_percent / 100 / 12
    emi_value = pmt(calc_monthly_rate, months, loan_amount)
    payments = read_payments(wb["Payments"], emi_value)

    rows = build_base_schedule(
        loan_amount=loan_amount,
        daily_rate_percent=daily_rate_percent,
        fee_annual_percent=fee_annual_percent,
        start_date=start_date,
        first_payment_date=first_payment_date,
        months=months,
        holidays=holidays,
    )

    apply_payments(rows, payments, daily_rate_percent)
    enforce_annuity_opening_chain(rows, daily_rate_percent)
    write_schedule_sheet(wb, rows)
    write_output_sheet(wb, rows)
    wb.save(args.output)
    print(f"Saved recalculated workbook to: {args.output}")


if __name__ == "__main__":
    main()
